from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView

from django.db import transaction
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated, BasePermission
from rest_framework.authentication import SessionAuthentication
from django.http import HttpResponse
from django.utils import timezone
from datetime import date

from .serializers import (
    UserSerializer,
    DashboardStateSerializer,
    LessonSerializer,
    StudentSerializer,
    RecordSerializer,
)
from .models import User, Lesson, Student, Record


class MeView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        serializer = UserSerializer(request.user)
        return Response(serializer.data)


class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = "dashboard.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        user: User = self.request.user
        ctx["can_edit"] = bool(
            user.is_superuser or user.role in [User.Roles.ADMIN, User.Roles.TEACHER]
        )
        # Expose whether the user is an admin (or superuser) so templates can hide destructive controls
        ctx["is_admin"] = bool(user.is_superuser or user.role == User.Roles.ADMIN)
        return ctx


class IsAdminOrTeacher(BasePermission):
    def has_permission(self, request, view):
        user: User = request.user
        if not user.is_authenticated:
            return False
        return user.role in [User.Roles.ADMIN, User.Roles.TEACHER] or user.is_superuser


class IsAdminOnly(BasePermission):
    def has_permission(self, request, view):
        user: User = request.user
        if not user.is_authenticated:
            return False
        return user.is_superuser or user.role == User.Roles.ADMIN


class DashboardStateView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Seed defaults on first run: 24 lessons and 30 students for each level
        if Lesson.objects.count() == 0:
            today = date.today()
            for i in range(24):
                # default sequential dates as today + i for convenience
                Lesson.objects.create(title=f"{i+1}-dars", order=i, date=today)
        # Ensure minimum rows per level
        # Keep compatibility: A2/B1/B2 seeded to 30 on first run; A1/C1/C2 ensure at least 1 row exists
        for level in [Student.Levels.A2, Student.Levels.B1, Student.Levels.B2]:
            if not Student.objects.filter(level=level).exists():
                Student.objects.bulk_create([Student(level=level, name="") for _ in range(30)])
        for level in [Student.Levels.A1, Student.Levels.C1, Student.Levels.C2]:
            if not Student.objects.filter(level=level).exists():
                Student.objects.create(level=level, name="")

        lessons = Lesson.objects.all()
        # Group students by all defined levels dynamically
        level_keys = [choice[0] for choice in Student.Levels.choices]
        students_by_level = { key: list(Student.objects.filter(level=key)) for key in level_keys }

        # Build records map: student_id -> lesson_id -> data
        records_map = {}
        recs = Record.objects.select_related('student', 'lesson').all()
        for r in recs:
            sid = str(r.student_id)
            lid = str(r.lesson_id)
            records_map.setdefault(sid, {})[lid] = {
                'attendance': r.attendance,
                'homework': r.homework,
                'extra': r.extra,
                'test_score': r.test_score,
            }

        data = {
            'lessons': LessonSerializer(lessons, many=True).data,
            'students': { key: StudentSerializer(students_by_level[key], many=True).data for key in students_by_level },
            'records': records_map,
        }
        return Response(DashboardStateSerializer(data).data)


class DashboardSaveView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    @transaction.atomic
    def post(self, request):
        payload = request.data
        # Expected payload: {
        #   records: { student_id: { lesson_id: {attendance<'P'|'E'|'A'>,homework,extra,test_score}}},
        #   students: [ {id,name,note} ],
        #   lessons: [ {id,date} ]
        # }
        records = payload.get('records', {})
        # Upsert records
        for sid, lessons in records.items():
            for lid, r in lessons.items():
                obj, _ = Record.objects.get_or_create(student_id=sid, lesson_id=lid)
                # Normalize attendance to valid choices: 'P','E','A'
                data = dict(r)
                att = data.get('attendance', 'A')
                if isinstance(att, bool):
                    att = 'P' if att else 'A'
                if att in (None, ''):
                    att = 'A'
                if att not in ('P','E','A'):
                    att = 'A'
                data['attendance'] = att
                ser = RecordSerializer(instance=obj, data=data, partial=True)
                ser.is_valid(raise_exception=True)
                ser.save()

        # Update students basic info if provided
        students_update = payload.get('students', [])
        for s in students_update:
            if 'id' in s:
                Student.objects.filter(id=s['id']).update(name=s.get('name', ''), note=s.get('note', ''))

        # Update lessons (dates)
        lessons_update = payload.get('lessons', [])
        for l in lessons_update:
            if 'id' in l:
                # allow blank/null date; parse ISO date string
                d = l.get('date')
                dt = None
                try:
                    if d:
                        dt = date.fromisoformat(d)
                except Exception:
                    dt = None
                Lesson.objects.filter(id=l['id']).update(date=dt)

        return Response({"status": "ok"})


class DashboardClearView(APIView):
    authentication_classes = [SessionAuthentication]
    # Only admins may clear all records
    permission_classes = [IsAuthenticated, IsAdminOnly]

    @transaction.atomic
    def post(self, request):
        Record.objects.all().delete()
        # Also clear student names and notes as requested
        Student.objects.all().update(name="", note="")
        return Response({"status": "cleared_all"})


class LessonAddView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def post(self, request):
        count = Lesson.objects.count()
        lesson = Lesson.objects.create(title=f"{count+1}-dars", order=count, date=date.today())
        return Response(LessonSerializer(lesson).data)


class LessonRemoveView(APIView):
    authentication_classes = [SessionAuthentication]
    # Only admins may remove lessons/columns
    permission_classes = [IsAuthenticated, IsAdminOnly]

    @transaction.atomic
    def post(self, request):
        # Enforce a minimum number of lessons kept
        MIN_LESSONS = 3
        if Lesson.objects.count() <= MIN_LESSONS:
            return Response({"status": "min_reached", "min": MIN_LESSONS})
        last = Lesson.objects.order_by('-order', '-id').first()
        if not last:
            return Response({"status": "noop"})
        Record.objects.filter(lesson=last).delete()
        last.delete()
        return Response({"status": "removed"})


class StudentAddView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated, IsAdminOrTeacher]

    def post(self, request):
        level = request.data.get('level')
        levels = [c[0] for c in Student.Levels.choices]
        if level not in levels:
            return Response({"error": "invalid_level", "levels": levels}, status=400)
        s = Student.objects.create(level=level, name="")
        return Response(StudentSerializer(s).data)


class StudentRemoveView(APIView):
    authentication_classes = [SessionAuthentication]
    # Only admins may remove student rows
    permission_classes = [IsAuthenticated, IsAdminOnly]

    def post(self, request):
        level = request.data.get('level')
        levels = [c[0] for c in Student.Levels.choices]
        if level not in levels:
            return Response({"error": "invalid_level", "levels": levels}, status=400)
        # Do not remove if it would drop below 1 row for A1/C1/C2, or 0 for others
        qs = Student.objects.filter(level=level).order_by('-id')
        if not qs.exists():
            return Response({"status": "noop"})
        # Keep at least 1 student per level
        if qs.count() <= 1:
            return Response({"status": "min_reached", "min": 1})
        stu = qs.first()
        stu.delete()
        return Response({"status": "removed"})


class DashboardExportView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Build an Excel (xlsx) file in-memory; fallback to CSV if openpyxl missing.
        lessons = list(Lesson.objects.all())
        students = list(Student.objects.all().order_by('level', 'id'))
        recs = Record.objects.select_related('student', 'lesson').all()
        records_map = {}
        for r in recs:
            records_map.setdefault(r.student_id, {})[r.lesson_id] = r

        # Determine which lessons have any meaningful data across all students
        def has_data(r: Record):
            if not r:
                return False
            if r.attendance in ('P', 'E'):
                return True
            if r.homework:
                return True
            if (r.extra or '').strip():
                return True
            if (r.test_score or 0) > 0:
                return True
            return False

        lessons_with_data = []
        for l in lessons:
            any_data = any(has_data(records_map.get(s.id, {}).get(l.id)) for s in students)
            if any_data:
                lessons_with_data.append(l)

        # Optionally filter out students that have no data at all
        students_with_data = []
        for s in students:
            any_data = any(has_data(records_map.get(s.id, {}).get(l.id)) for l in lessons_with_data)
            if any_data:
                students_with_data.append(s)

        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            wb = Workbook()
            ws = wb.active
            ws.title = 'Dashboard'

            # Headers
            headers = ['#', 'Ism Familiya', 'Daraja', 'Qo\'shimcha izoh']
            for idx, h in enumerate(headers, start=1):
                ws.cell(row=1, column=idx, value=h)
            col = len(headers) + 1
            for l in lessons_with_data:
                ws.cell(row=1, column=col, value=f"{l.title}")
                col += 1

            # Rows
            row_idx = 2
            for i, s in enumerate(students_with_data, start=1):
                ws.cell(row=row_idx, column=1, value=i)
                ws.cell(row=row_idx, column=2, value=s.name)
                ws.cell(row=row_idx, column=3, value=s.level)
                ws.cell(row=row_idx, column=4, value=s.note)
                col = 5
                for l in lessons_with_data:
                    rec = records_map.get(s.id, {}).get(l.id)
                    att = rec.attendance if rec else ''
                    symbol = {'P': '+', 'E': '−', 'A': ''}.get(att, '')  # hide plain '×' to avoid clutter
                    ws.cell(row=row_idx, column=col, value=symbol)
                    col += 1
                row_idx += 1

            from io import BytesIO
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            resp = HttpResponse(
                buf.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            resp['Content-Disposition'] = 'attachment; filename="dashboard.xlsx"'
            return resp
        except Exception:
            # Fallback to CSV
            import csv
            from io import StringIO
            sio = StringIO()
            writer = csv.writer(sio)
            writer.writerow(['#', 'Ism Familiya', 'Daraja', 'Izoh'] + [l.title for l in lessons_with_data])
            for i, s in enumerate(students_with_data, start=1):
                row = [i, s.name, s.level, s.note]
                for l in lessons_with_data:
                    rec = records_map.get(s.id, {}).get(l.id)
                    att = rec.attendance if rec else ''
                    row.append({'P': '+', 'E': '−', 'A': ''}.get(att, ''))
                writer.writerow(row)
            resp = HttpResponse(sio.getvalue(), content_type='text/csv')
            resp['Content-Disposition'] = 'attachment; filename="dashboard.csv"'
            return resp
